from selenium import webdriver
import time
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import wareki
import mojimoji

def con_chrome(path):
    wb = openpyxl.load_workbook(path) #エクセルファイルの読み込み
    ws = wb.active #エクセルシートの読み込み
    max_num=wb['雇用保険被保険者資格喪失届（連記式）'].max_row
    #Chromeを操作
    #合意画面
    ChromeOptions = webdriver.ChromeOptions()
    ChromeOptions.add_experimental_option('excludeSwitches', ['enable-logging'])
    ChromeOptions.add_experimental_option('detach', True)
    driver = webdriver.Chrome(ChromeDriverManager().install(),options=ChromeOptions)
    driver.get("https://hoken.hellowork.mhlw.go.jp/assist/001000.do?screenId=001000&action=koyohohiSoshitsuLink")
    driver.find_element_by_id('ID_chkDoi1').click()
    driver.find_element_by_id('ID_inputPrintButton').click()
    #入力画面
    for i in range(2,max_num+1):
        #被保険者番号入力
        if ws.cell(i,7).value!=None:
            driver.execute_script('document.getElementById("ID_txtHhksNo1").value="'+ str(ws.cell(i,7).value) +'";')
        if ws.cell(i,8).value!=None:
            driver.execute_script('document.getElementById("ID_txtHhksNo2").value="'+ str(ws.cell(i,8).value) +'";')
        if ws.cell(i,9).value!=None:
            driver.execute_script('document.getElementById("ID_txtHhksNo3").value="'+ str(ws.cell(i,9).value) +'";')

        #事業所番号入力
        if ws.cell(i,10).value!=None:
            driver.execute_script('document.getElementById("ID_txtJgshNo1").value="'+ str(ws.cell(i,10).value) +'";')
        if ws.cell(i,11).value!=None:
            driver.execute_script('document.getElementById("ID_txtJgshNo2").value="'+ str(ws.cell(i,11).value) +'";')
        if ws.cell(i,12).value!=None:
            driver.execute_script('document.getElementById("ID_txtJgshNo3").value="'+ str(ws.cell(i,12).value) +'";')

        #資格取得年月日入力
        if ws.cell(i,13).value!=None:
            YMD=ws.cell(i,13).value.date()
            seireki,year=wareki.wareki_conv(YMD)
            driver.execute_script('document.getElementById("ID_skkuStkYmdGG").value="'+ wareki.conv_num(seireki) +'";')
            driver.execute_script('document.getElementById("ID_skkuStkYmdYY").value="'+ year +'";')
            driver.execute_script('document.getElementById("ID_skkuStkYmdMM").value="'+ str(YMD.month).zfill(2) +'";')
            driver.execute_script('document.getElementById("ID_skkuStkYmdDD").value="'+ str(YMD.day).zfill(2) +'";')

        #離職等年月日入力
        if ws.cell(i,14).value!=None:
            YMD=ws.cell(i,14).value.date()
            seireki,year=wareki.wareki_conv(YMD)
            driver.execute_script('document.getElementById("ID_riskYmdGG").value="'+ wareki.conv_num(seireki) +'";')
            driver.execute_script('document.getElementById("ID_riskYmdYY").value="'+ year +'";')
            driver.execute_script('document.getElementById("ID_riskYmdMM").value="'+ str(YMD.month).zfill(2) +'";')
            driver.execute_script('document.getElementById("ID_riskYmdDD").value="'+ str(YMD.day).zfill(2) +'";')
        #6．喪失原因
        if ws.cell(i,15).value!=None:
            driver.find_element_by_id('ID_rdoSoshitsuGenin'+str(ws.cell(i,15).value)).click()
        #7．離職票交付希望
        #driver.find_element_by_id('ID_rdoRiskHyoKofuKibo'+str(ws.cell(i,17).value)).click()
        #8.１週間の所定労働時間
        if ws.cell(i,26).value!=None:
            driver.execute_script('document.getElementById("ID_txt1ShuukanNoShoteiRodoJnJn").value="'+ str(ws.cell(i,26).value) +'";')
        if ws.cell(i,27).value!=None:
            driver.execute_script('document.getElementById("ID_txt1ShuukanNoShoteiRodoJnFun").value="'+ str(ws.cell(i,27).value) +'";')
        #9．補充採用予定の有無
        if ws.cell(i,17).value!=None:
            driver.find_element_by_id('ID_rdoHojuSaiyoYoteiNoUmu'+str(ws.cell(i,17).value)).click()
        #フリガナ（カタカナ）10．新氏名（漢字)
        kana=ws.cell(i,5).value
        kanji=ws.cell(i,6).value
        if kana!=None:
            driver.execute_script('document.getElementById("ID_txtFuriganaKatakana").value="'+ kana +'";')
        if kanji!=None:
            driver.execute_script('document.getElementById("ID_txtNewNameKnj").value="'+ kanji +'";')
        #（フリガナ）20．被保険者氏名（漢字)
        kana=ws.cell(i,4).value
        kanji=ws.cell(i,3).value
        if kana!=None:
            driver.execute_script('document.getElementById("ID_txtHhksNameKana").value="'+ mojimoji.han_to_zen(kana) +'";')
        if kanji!=None:
            driver.execute_script('document.getElementById("ID_txtHhksNameKnj").value="'+ kanji +'";')
        #21．性別
        if ws.cell(i,19).value!=None:
            driver.find_element_by_id('ID_rdoSeibetsu'+str(ws.cell(i,19).value)).click()
        #22．生年月日
        if ws.cell(i,21).value!=None:
            YMD=ws.cell(i,21).value.date()
            seireki,year=wareki.wareki_conv(YMD)
            driver.execute_script('document.getElementById("ID_seinenGG").value="'+ wareki.conv_num(seireki) +'";')
            driver.execute_script('document.getElementById("ID_seinenYY").value="'+ year +'";')
            driver.execute_script('document.getElementById("ID_seinenMM").value="'+ str(YMD.month).zfill(2) +'";')
            driver.execute_script('document.getElementById("ID_seinenDD").value="'+ str(YMD.day).zfill(2) +'";')
        #23．被保険者の住所又は居所
        if ws.cell(i,22).value!=None:
            jusho=ws.cell(i,22).value
            driver.execute_script('document.getElementById("ID_txtHhksJusho").value="'+ jusho +'";')
        #24．事業所名称
        if ws.cell(i,23).value!=None:
            jgyosho=ws.cell(i,23).value
            driver.execute_script('document.getElementById("ID_txtJgshMeisho").value="'+ jgyosho +'";')
        #25．氏名変更年月日
        if ws.cell(i,24).value!=None:
            YMD=ws.cell(i,24).value.date()
            seireki,year=wareki.wareki_conv(YMD)
            driver.execute_script('document.getElementById("ID_nameChangeYmdGG").value="'+ wareki.conv_num(seireki) +'";')
            driver.execute_script('document.getElementById("ID_nameChangeYmdYY").value="'+ year +'";')
            driver.execute_script('document.getElementById("ID_nameChangeYmdMM").value="'+ str(YMD.month).zfill(2) +'";')
            driver.execute_script('document.getElementById("ID_nameChangeYmdDD").value="'+ str(YMD.day).zfill(2) +'";')
        #26．被保険者でなくなったことの原因
        if ws.cell(i,25).value!=None:
            riyuu=ws.cell(i,25).value
            driver.execute_script('document.getElementById("ID_txtHhksChangeCause").value="'+ riyuu +'";')
        #申請先
        if ws.cell(i,28).value!=None:
            driver.execute_script('document.getElementById("ID_txtShinseiSaki").value="'+ ws.cell(i,28).value +'";')
        #pdfダウンロード
        driver.find_element_by_id('ID_commonDownload').click()
        if i!=max_num:
            driver.find_element_by_id('ID_continueMakeButton').click()
        time.sleep(5)
    #driver.quit()